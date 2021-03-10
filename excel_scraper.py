from openpyxl import load_workbook

# https://openpyxl.readthedocs.io/en/stable/tutorial.html

wb = load_workbook("Testbig.xlsx")
ws = wb.active

col = ws["A"]

for cell in col:
    if isinstance(cell.value, int):
        print(cell)
        x = cell.value + 1
        print(cell.row)
        ws.cell(row=(cell.row), column=(cell.column + 1), value=x)

wb.save("Testbig.xlsx")




