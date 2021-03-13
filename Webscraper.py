from requests_html import HTMLSession
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys

# Test BIG-numbers
# BIG geldig maar geen HA
#BIG = 99924036101 Franken
# BIG huisarts
#BIG = 19059287401 Jansen
# BIG ongeldig
#BIG = 59919419602

# We use openpyxl to get and put data in an excel sheet
# The program wil fail if the excel file is not closed.
# So I created a custom message for that situation with try/except.
try:
    wb = load_workbook("Testbig.xlsx")
except:
    print("ERROR: Something went wrong loading the excel file. Did your remember to close the excel file before using?")
    sys.exit()
ws = wb.active

# We presume that the BIG-numbers are stored in column A of the excel sheet
col = ws["A"]

# We loop trough all BIG-numbers in the column.
# Openpyxl makes sure the looping stops at the first empty cell in the column
for cell in col:
    # We check whether the cell contains a integer-number (skipping the header of the column)
    if isinstance(cell.value, int):
        # We create the URL to load using the BIG-number from the cell in the column
        url = ("https://zoeken.bigregister.nl/bignummer/%s" % str(cell.value))

        # This code block uses HTMLsession
        # https://requests.readthedocs.io/projects/requests-html/en/latest/
        session = HTMLSession()
        response = session.get(url)
        # With adding "sleep" we wait a short while while the script is executed
        # than after waiting we get the website as loaded
        # without sleep it renders too fast and we get the website before the script is executed
        response.html.render(sleep=0.1)

        # We place the rendered website in a variable called "text"
        text = response.html.text

        ### ----------------------- Checking BIG ----------------------------- ###

        # We use parsing to search the text
        # only when a BIG-number is true the word "Naam" occurs
        # only when a true Big-number belongs to a huisarts the word "Huisartsgeneeskunde" occurs
        # When the word does not occur it will produce "-1"
        Check_BIG_Naam = text.find("Naam")
        Check_BIG_Huisartsgeneeskunde = text.find("Huisartsgeneeskunde")

        print(cell.value)
        if Check_BIG_Huisartsgeneeskunde > -1:
            Outcome = "BIG is gekoppeld aan huisarts"
            Cell_fill_color = "99CC00" #Green
        elif Check_BIG_Naam > -1:
            Outcome = "BIG is geldig, maar geen huisarts"
            Cell_fill_color = "FF9900" #Orange
        else:
            Outcome = "BIG levert geen resultaat op"
            Cell_fill_color = "FF0000" #Red

        print(Outcome)
        # We store the outcome in the column (+2) the cell right of the BIG-number
        Two_cells_to_the_right = ws.cell(row=(cell.row), column=(cell.column + 2))
        Two_cells_to_the_right.value = Outcome
        # We change the color of the cell depending on the outcome
        Two_cells_to_the_right.fill = PatternFill("solid", start_color=Cell_fill_color)

        ### ----------------------- Checking last name ----------------------------- ###

        # We take the name of the doctor from the excel (2nd row)
        # With parsing we check whether this name is also found when searching for the BIG-number
        Name_in_excel = ws.cell(row=(cell.row), column=(cell.column + 1)).value

        # We search the text for the name of the doctor in the excel
        # This search will give the "location" of the Name (around 230)
        # if the name is not found it will return -1
        Name_location = text.find(Name_in_excel)

        # To be sure the name is exactly the same we check whether there is a space in front of it
        # we use the location of where we found the name -1 place
        # to prevent: search --> Pol and get positive on Repol
        Space_before_name = text.find(" ", (Name_location - 1), Name_location)

        # We also check if there is a "G" right after the name. This is the first letter on the next row.
        # we use the location of the name + length of the name + 1
        # To prevent: search --> Frank and get positive on Franken
        Letters_in_name = len(Name_in_excel)
        G_after_name = text.find("G", (Name_location + Letters_in_name + 1),
                        (Name_location + Letters_in_name + 2))

        # If one of these 3 checks is -1 we have a mismatch in the name
        if Name_location < 0 or Space_before_name < 0 or G_after_name <0:
            print("Naam komt niet overeen")
            Outcome = "Naam komt niet overeen"
            Cell_fill_color = "FF0000"  # Red
        else:
            print("Naam komt overeen")
            Outcome = "Naam komt overeen"
            Cell_fill_color = "99CC00"  # Green

        # We store the outcome in the column (+2) the cell right of the Name
        Two_cells_to_the_right = ws.cell(row=(cell.row), column=(cell.column + 3))
        Two_cells_to_the_right.value = Outcome
        # We change the color of the cell depending on the outcome
        Two_cells_to_the_right.fill = PatternFill("solid", start_color=Cell_fill_color)

        print("------------------------------------")

# We save the changes to the excel
wb.save("Testbig.xlsx")


